import os
import re
import html
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles.colors import Color

REPO_ROOT = os.path.join(os.path.dirname(__file__), "..")
DOCS_DIR = os.path.join(REPO_ROOT, "docs")

XLSX_PATH = os.path.join(DOCS_DIR, "Schedule 2026.xlsx")
OUT_HTML = os.path.join(DOCS_DIR, "Schedule.html")


# ----------------------------
# Theme color support helpers
# ----------------------------
def _hex_to_rgb(hex6: str):
    hex6 = hex6.strip().lstrip("#")
    return int(hex6[0:2], 16), int(hex6[2:4], 16), int(hex6[4:6], 16)

def _rgb_to_hex(r: int, g: int, b: int):
    return f"#{r:02X}{g:02X}{b:02X}"

def _apply_tint_to_rgb(r: int, g: int, b: int, tint: float):
    tint = float(tint)
    def adj(c):
        if tint < 0:
            return int(round(c * (1.0 + tint)))
        else:
            return int(round(c + (255 - c) * tint))
    r2, g2, b2 = adj(r), adj(g), adj(b)
    return max(0, min(255, r2)), max(0, min(255, g2)), max(0, min(255, b2))

def _get_theme_palette_hex(wb) -> list:
    fallback = [
        "FFFFFF","000000","EEECE1","1F497D",
        "4F81BD","C0504D","9BBB59","8064A2",
        "4BACC6","F79646","0000FF","800080"
    ]
    try:
        theme = wb.theme
        if theme is None:
            return fallback
        cs = theme.themeElements.clrScheme
        names = ["lt1","dk1","lt2","dk2","accent1","accent2","accent3","accent4","accent5","accent6","hlink","folHlink"]
        out = []
        for nm in names:
            cobj = getattr(cs, nm, None)
            if cobj is None:
                out.append(fallback[len(out)])
                continue
            val = None
            if getattr(cobj, "srgbClr", None) is not None:
                val = getattr(cobj.srgbClr, "val", None)
            if val is None and getattr(cobj, "sysClr", None) is not None:
                val = getattr(cobj.sysClr, "lastClr", None) or getattr(cobj.sysClr, "val", None)
            if not val:
                out.append(fallback[len(out)])
            else:
                v = str(val).strip().lstrip("#")
                out.append(v.upper() if len(v) == 6 else fallback[len(out)])
        return out
    except Exception:
        return fallback

def _css_color_from_openpyxl_color(c: Color, theme_palette_hex: list):
    if c is None:
        return None

    rgb = getattr(c, "rgb", None)
    if rgb:
        rgb = str(rgb).strip()
        if len(rgb) == 8:
            rgb = rgb[2:]
        if len(rgb) == 6:
            return f"#{rgb.upper()}"
        return None

    theme_idx = getattr(c, "theme", None)
    if theme_idx is not None:
        try:
            idx = int(theme_idx)
        except Exception:
            return None
        if 0 <= idx < len(theme_palette_hex):
            base_hex = theme_palette_hex[idx]
            r, g, b = _hex_to_rgb(base_hex)
            tint = getattr(c, "tint", None)
            if tint is not None:
                r, g, b = _apply_tint_to_rgb(r, g, b, float(tint))
            return _rgb_to_hex(r, g, b)

    return None


# ----------------------------
# Excel -> HTML styling
# ----------------------------
def _cell_style_to_css(cell, theme_palette_hex):
    styles = []

    f = cell.font
    if f is not None:
        if f.bold:
            styles.append("font-weight:700")
        if f.italic:
            styles.append("font-style:italic")
        if f.underline:
            styles.append("text-decoration:underline")
        if f.color is not None:
            col = _css_color_from_openpyxl_color(f.color, theme_palette_hex)
            if col:
                styles.append(f"color:{col}")

    fill = cell.fill
    if fill is not None and getattr(fill, "patternType", None) == "solid":
        fg = getattr(fill, "fgColor", None)
        col = _css_color_from_openpyxl_color(fg, theme_palette_hex)
        if col:
            styles.append(f"background:{col}")

    a = cell.alignment
    if a is not None:
        if a.horizontal:
            styles.append(f"text-align:{a.horizontal}")
        if a.vertical:
            styles.append(f"vertical-align:{a.vertical}")

    return ";".join(styles)

def _escape_cell_value(v):
    if v is None:
        return "&nbsp;"
    s = html.escape(str(v))
    s = s.replace("\n", "<br>")
    return s if s.strip() else "&nbsp;"


def _row_is_blank(ws, r, max_col):
    for c in range(1, max_col + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip() != "":
            return False
    return True


def main():
    if not os.path.isfile(XLSX_PATH):
        raise SystemExit(f"ERROR: Missing file: {XLSX_PATH}")

    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    theme_palette_hex = _get_theme_palette_hex(wb)

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # Find the first blank row that separates schedule table from notes (your screenshot shows this pattern)
    # If we can't find one, treat whole sheet as schedule.
    split_row = None
    for r in range(1, max_row + 1):
        if _row_is_blank(ws, r, max_col):
            split_row = r
            break

    schedule_end = (split_row - 1) if split_row else max_row
    notes_start = (split_row + 1) if split_row else None

    # merged-cell maps
    merged_top_left = {}
    merged_covered = set()
    for m in ws.merged_cells.ranges:
        rs = m.max_row - m.min_row + 1
        cs = m.max_col - m.min_col + 1
        merged_top_left[(m.min_row, m.min_col)] = (rs, cs)
        for rr in range(m.min_row, m.max_row + 1):
            for cc in range(m.min_col, m.max_col + 1):
                if (rr, cc) != (m.min_row, m.min_col):
                    merged_covered.add((rr, cc))

    # If row 1 is a "title row" with only one non-empty cell and it's NOT merged across,
    # force it to span the full width so it becomes the header bar.
    def row1_title_colspan_override():
        nonempty = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=1, column=c).value
            if v is not None and str(v).strip() != "":
                nonempty.append(c)
        if len(nonempty) == 1:
            c = nonempty[0]
            if (1, c) not in merged_top_left:
                return c
        return None

    title_col = row1_title_colspan_override()

    updated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with open(OUT_HTML, "w", encoding="utf-8") as f:
        f.write("<!doctype html><html><head><meta charset='utf-8'>")
        f.write("<title>Schedule</title>")
        # Compact CSS: single-line cells to keep height down; responsive width.
        f.write(
            "<style>"
            "html,body{margin:0;padding:0}"
            "body{font-family:Calibri,Arial;background:#ffffff}"
            ".wrap{max-width:99vw;margin:8px auto;border:3px solid #000;background:#FFFFCC;padding:8px;box-sizing:border-box}"
            ".meta{font-size:10pt;margin:0 0 8px 0}"
            "table{border-collapse:collapse;width:100%;table-layout:fixed;background:#ffffff}"
            "th,td{border:1px solid #000;padding:2px 4px;font-size:10pt;line-height:1.05;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}"
            "th{background:#c0c0c0}"
            ".titlecell{font-size:18pt;font-weight:700;text-align:center;background:#c0c0c0;padding:10px 6px}"
            ".notes th,.notes td{font-size:10pt;white-space:normal;overflow:visible;text-overflow:clip}"
            "</style>"
        )
        f.write("</head><body><div class='wrap'>")
        f.write(f"<div class='meta'><b>Last updated:</b> {html.escape(updated)}</div>")

        # -------- Schedule table (fit one screen) --------
        f.write("<table class='schedule'>")

        for r in range(1, schedule_end + 1):
            f.write("<tr>")
            for c in range(1, max_col + 1):
                if (r, c) in merged_covered:
                    continue

                cell = ws.cell(row=r, column=c)

                # First two rows are headers per your sheet
                tag = "th" if r <= 2 else "td"

                attrs = []
                css = _cell_style_to_css(cell, theme_palette_hex)

                # Title row: force colspan across all columns
                if r == 1 and title_col is not None and c == title_col:
                    attrs.append(f"colspan='{max_col}'")
                    attrs.append("class='titlecell'")
                    # ignore any Excel alignment for this specific cell; the CSS will handle it
                    css = ""

                    # Also skip the rest of the row’s cells (they’ll be empty anyway)
                    val = _escape_cell_value(cell.value)
                    f.write(f"<td {' '.join(attrs)}>{val}</td>")
                    break

                # Honor real merges elsewhere
                if (r, c) in merged_top_left:
                    rs, cs = merged_top_left[(r, c)]
                    if rs > 1:
                        attrs.append(f"rowspan='{rs}'")
                    if cs > 1:
                        attrs.append(f"colspan='{cs}'")

                if css:
                    attrs.append(f"style='{css}'")

                val = _escape_cell_value(cell.value)
                f.write(f"<{tag} {' '.join(attrs)}>{val}</{tag}>")

            f.write("</tr>")

        f.write("</table>")

        # -------- Notes section (can extend downward) --------
        if notes_start and notes_start <= max_row:
            f.write("<div style='height:10px'></div>")
            f.write("<table class='notes'>")
            for r in range(notes_start, max_row + 1):
                f.write("<tr>")
                for c in range(1, max_col + 1):
                    if (r, c) in merged_covered:
                        continue
                    cell = ws.cell(row=r, column=c)

                    attrs = []
                    if (r, c) in merged_top_left:
                        rs, cs = merged_top_left[(r, c)]
                        if rs > 1:
                            attrs.append(f"rowspan='{rs}'")
                        if cs > 1:
                            attrs.append(f"colspan='{cs}'")

                    css = _cell_style_to_css(cell, theme_palette_hex)
                    if css:
                        attrs.append(f"style='{css}'")

                    val = _escape_cell_value(cell.value)
                    f.write(f"<td {' '.join(attrs)}>{val}</td>")
                f.write("</tr>")
            f.write("</table>")

        f.write("<div style='margin-top:10px;font-size:10pt;'>")
        f.write(f"<a href='{html.escape(os.path.basename(XLSX_PATH))}'>Download the Excel version</a>")
        f.write("</div>")

        f.write("</div></body></html>")

    print(f"Wrote: {OUT_HTML}")


if __name__ == "__main__":
    main()
