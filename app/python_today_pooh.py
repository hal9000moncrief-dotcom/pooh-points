import sys
import time
import random
import re
import requests
import os
import html
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

BASE = "https://site.api.espn.com/apis/site/v2/sports/basketball/mens-college-basketball"
SEC_TEAMS_HTML = "https://www.espn.com/mens-college-basketball/teams/_/group/23"

DRAFT_XLSX = os.path.join(os.path.dirname(__file__), "ByCoach.xlsx")   # must be in same folder

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0 Safari/537.36",
    "Accept": "application/json,text/plain,*/*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.espn.com/",
    "Connection": "keep-alive",
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)

BASE_DELAY = 0.25
JITTER = 0.25
MAX_RETRIES = 6
TIMEOUT = 30


# ----------------------------
# UTIL
# ----------------------------
def polite_sleep():
    time.sleep(BASE_DELAY + random.random() * JITTER)

def get_text(url: str) -> str:
    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = SESSION.get(url, timeout=TIMEOUT)
            r.raise_for_status()
            polite_sleep()
            return r.text
        except Exception as e:
            last_err = e
            time.sleep((0.7 ** attempt) + random.random() * 0.7)
    raise RuntimeError(f"Failed after retries: {url}\nLast error: {last_err}")

def get_json(url: str) -> dict:
    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = SESSION.get(url, timeout=TIMEOUT)
            r.raise_for_status()
            polite_sleep()
            return r.json()
        except Exception as e:
            last_err = e
            time.sleep((0.7 ** attempt) + random.random() * 0.7)
    raise RuntimeError(f"Failed after retries: {url}\nLast error: {last_err}")

def safe_int(v) -> int:
    try:
        return int(str(v).strip())
    except:
        return 0

def parse_made_attempt(s: str) -> Tuple[int, int]:
    try:
        a, b = str(s).split("-")
        return int(a), int(b)
    except:
        return 0, 0

def to_minutes(v) -> float:
    if v is None:
        return 0.0
    s = str(v).strip()
    if not s or s == "--":
        return 0.0
    if ":" in s:
        try:
            mm, ss = s.split(":")
            return int(mm) + int(ss) / 60.0
        except:
            return 0.0
    try:
        return float(s)
    except:
        return 0.0

def norm_name(name: str) -> str:
    s = (name or "").lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\b(jr|sr|ii|iii|iv)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def compute_pooh(values: List[str], labels: List[str]) -> Optional[dict]:
    # Pooh = PTS + REB + AST + STL + BLK - missedFG - missedFT - TO
    if not labels or not values:
        return None

    def idx(name: str) -> Optional[int]:
        try:
            return labels.index(name)
        except ValueError:
            return None

    i_min = idx("MIN")
    i_fg  = idx("FG")
    i_ft  = idx("FT")
    i_reb = idx("REB")
    i_ast = idx("AST")
    i_stl = idx("STL")
    i_blk = idx("BLK")
    i_to  = idx("TO")
    i_pts = idx("PTS")

    required = [i_min, i_fg, i_ft, i_reb, i_ast, i_stl, i_blk, i_to, i_pts]
    if any(x is None for x in required):
        return None

    max_i = max(required)
    if len(values) <= max_i:
        return None

    mins = to_minutes(values[i_min])
    fgm, fga = parse_made_attempt(values[i_fg])
    ftm, fta = parse_made_attempt(values[i_ft])

    missed_fg = max(0, fga - fgm)
    missed_ft = max(0, fta - ftm)

    pts = safe_int(values[i_pts])
    reb = safe_int(values[i_reb])
    ast = safe_int(values[i_ast])
    stl = safe_int(values[i_stl])
    blk = safe_int(values[i_blk])
    tov = safe_int(values[i_to])

    # Skip truly blank/DNP rows
    if mins == 0 and pts == 0 and reb == 0 and ast == 0 and stl == 0 and blk == 0 and tov == 0 and fga == 0 and fta == 0:
        return None

    pooh = (pts + reb + ast + stl + blk) - (missed_fg + missed_ft + tov)

    return {
        "MIN": mins,
        "PTS": pts, "REB": reb, "AST": ast, "STL": stl, "BLK": blk, "TO": tov,
        "POOH": pooh,
    }


# ----------------------------
# DRAFT BOARD
# ----------------------------
def load_draft_board(xlsx_path: str) -> Tuple[Dict[str, dict], List[str]]:
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = {}
    for cell in ws[1]:
        h = str(cell.value).strip() if cell.value is not None else ""
        headers[h.lower()] = cell.column

    if "name" not in headers or "owner" not in headers:
        raise RuntimeError("ByCoach.xlsx must have columns: Name, Owner (and optional Started).")

    col_name = headers["name"]
    col_owner = headers["owner"]
    col_started = headers.get("started", None)

    draft_map: Dict[str, dict] = {}
    owner_order: List[str] = []

    for r in range(2, ws.max_row + 1):
        nm = ws.cell(row=r, column=col_name).value
        ow = ws.cell(row=r, column=col_owner).value
        st = ws.cell(row=r, column=col_started).value if col_started else None

        name = str(nm).strip() if nm else ""
        owner = str(ow).strip() if ow else ""
        started_raw = str(st).strip().lower() if st is not None else ""
        started = started_raw in ("yes", "y", "true", "1")

        if not name:
            continue
        key = norm_name(name)
        if not owner:
            owner = "Undrafted"

        if key not in draft_map:
            draft_map[key] = {"owner": owner, "started": started, "raw_name": name}

        if owner and owner != "Undrafted" and owner not in owner_order:
            owner_order.append(owner)

    return draft_map, owner_order


# ----------------------------
# SEC + SCOREBOARD
# ----------------------------
def get_sec_team_ids() -> set:
    html = get_text(SEC_TEAMS_HTML)
    soup = BeautifulSoup(html, "lxml")
    team_ids = set()
    for a in soup.select('a[href*="/mens-college-basketball/team/_/id/"]'):
        href = a.get("href", "")
        m = re.search(r"/id/(\d+)", href)
        if m:
            team_ids.add(m.group(1))
    if len(team_ids) != 16:
        print(f"WARNING: SEC team IDs parsed = {len(team_ids)} (expected 16)")
    return team_ids

def get_today_events(date_yyyymmdd: str) -> List[dict]:
    url = f"{BASE}/scoreboard?dates={date_yyyymmdd}&groups=50&limit=500"
    data = get_json(url)
    return data.get("events", [])

def extract_event_header(e: dict) -> dict:
    comps = e.get("competitions") or []
    comp = comps[0] if comps else {}

    status_obj = comp.get("status", {}) or {}
    status_type = status_obj.get("type", {}) or {}
    detail = status_type.get("detail") or status_type.get("description") or status_type.get("name") or "Unknown"

    competitors = comp.get("competitors") or []
    ha = {}
    for c in competitors:
        ha_key = c.get("homeAway") or ""
        team = c.get("team", {}) or {}
        ha[ha_key] = {
            "id": str(team.get("id") or ""),
            "abbr": team.get("abbreviation") or "",
            "name": team.get("displayName") or team.get("shortDisplayName") or "",
            "score": safe_int(c.get("score")),
        }

    return {"status": detail, "home": ha.get("home", {}), "away": ha.get("away", {})}

def is_sec_involved(event_obj: dict, sec_ids: set) -> bool:
    hdr = extract_event_header(event_obj)
    return (hdr["home"].get("id", "") in sec_ids) or (hdr["away"].get("id", "") in sec_ids)


# ----------------------------
# BOXSCORE PARSE
# ----------------------------
def iter_athlete_rows(stat_group: dict) -> List[dict]:
    rows = []
    for key in ("athletes", "bench", "reserves"):
        v = stat_group.get(key)
        if isinstance(v, list):
            rows.extend(v)
    if not rows and isinstance(stat_group.get("athletes"), list):
        rows = stat_group["athletes"]
    return rows

def get_boxscore_players(event_id: str) -> List[dict]:
    url = f"{BASE}/summary?event={event_id}"
    data = get_json(url)

    box = data.get("boxscore") or {}
    players_sections = box.get("players") or []
    if not players_sections:
        return []

    out = []
    for ps in players_sections:
        team = ps.get("team", {}) or {}
        tabbr = team.get("abbreviation") or ""

        seen = set()
        for stat_group in ps.get("statistics") or []:
            labels = stat_group.get("labels") or []
            if not labels:
                continue

            for ath in iter_athlete_rows(stat_group):
                athlete = ath.get("athlete", {}) or {}
                aid = str(athlete.get("id") or "")
                pname = athlete.get("displayName") or athlete.get("shortName") or athlete.get("fullName") or "Unknown"
                values = ath.get("stats") or []

                if aid and aid in seen:
                    continue

                line = compute_pooh(values, labels)
                if not line:
                    continue

                if aid:
                    seen.add(aid)

                out.append({
                    "team": tabbr,
                    "player": pname,
                    "pooh": line["POOH"],
                    "pts": line["PTS"],
                    "reb": line["REB"],
                    "ast": line["AST"],
                    "stl": line["STL"],
                    "blk": line["BLK"],
                    "to":  line["TO"],
                    "min": line["MIN"],
                })

    return out


# ----------------------------
# XLSX OUTPUT
# ----------------------------
def autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

def write_xlsx(players_rows: List[dict], owner_totals_rows: List[dict], out_path: str):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Players"

    headers1 = ["date","owner","started_today","player","team","game","status","pooh","pts","reb","ast","stl","blk","to","min"]
    ws1.append(headers1)
    for c in range(1, len(headers1) + 1):
        ws1.cell(row=1, column=c).font = Font(bold=True)
        ws1.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    for r in players_rows:
        ws1.append([r.get(h, "") for h in headers1])

    ws2 = wb.create_sheet("OwnerTotals")
    headers2 = ["owner", "starter_pooh_total", "starters_count_so_far"]
    ws2.append(headers2)
    for c in range(1, len(headers2) + 1):
        ws2.cell(row=1, column=c).font = Font(bold=True)
        ws2.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    for r in owner_totals_rows:
        ws2.append([r["owner"], r["starter_pooh_total"], r["starters_count_so_far"]])

    autosize_columns(ws1)
    autosize_columns(ws2)

    wb.save(out_path)

def write_html_tables(players_rows, owner_totals_rows, out_players_html, out_owners_html, date_str):
    def esc(x):
        return html.escape("" if x is None else str(x))

    players_cols = ["owner","started_today","player","team","game","status","pooh","pts","reb","ast","stl","blk","to","min"]

    # Players page
    with open(out_players_html, "w", encoding="utf-8") as f:
        f.write("<!doctype html><html><head><meta charset='utf-8'>")
        f.write(f"<title>SEC Pooh Points — {esc(date_str)}</title>")
        f.write("<style>body{font-family:Arial}table{border-collapse:collapse;font-size:14px}"
                "th,td{border:1px solid #ccc;padding:4px 6px}th{background:#eee}"
                ".start{font-weight:bold}</style>")
        f.write("</head><body>")
        f.write(f"<h2>SEC Pooh Points — {esc(date_str)}</h2>")
        f.write("<table><thead><tr>")
        for c in players_cols:
            f.write(f"<th>{esc(c)}</th>")
        f.write("</tr></thead><tbody>")
        for r in players_rows:
            cls = " class='start'" if r.get("started_today") == "Yes" else ""
            f.write("<tr>")
            for c in players_cols:
                f.write(f"<td{cls}>{esc(r.get(c,''))}</td>")
            f.write("</tr>")
        f.write("</tbody></table></body></html>")

    # Owners page
    with open(out_owners_html, "w", encoding="utf-8") as f:
        f.write("<!doctype html><html><head><meta charset='utf-8'>")
        f.write(f"<title>Owner Starters Total — {esc(date_str)}</title>")
        f.write("<style>body{font-family:Arial}table{border-collapse:collapse;font-size:14px}"
                "th,td{border:1px solid #ccc;padding:4px 6px}th{background:#eee}</style>")
        f.write("</head><body>")
        f.write(f"<h2>Owner Starters Total — {esc(date_str)}</h2>")
        f.write("<table><thead><tr><th>Owner</th><th>Starter Pooh Total</th><th>Starters Count So Far</th></tr></thead><tbody>")
        for r in owner_totals_rows:
            f.write(f"<tr><td>{esc(r['owner'])}</td><td>{esc(r['starter_pooh_total'])}</td><td>{esc(r['starters_count_so_far'])}</td></tr>")
        f.write("</tbody></table></body></html>")
# ----------------------------
# MAIN
# ----------------------------
def main():
    if len(sys.argv) >= 2:
        date_yyyymmdd = sys.argv[1].strip()
    else:
        date_yyyymmdd = datetime.now().strftime("%Y%m%d")

    draft_map, owner_order = load_draft_board(DRAFT_XLSX)

    sec_ids = get_sec_team_ids()
    events = get_today_events(date_yyyymmdd)
    sec_events = [e for e in events if is_sec_involved(e, sec_ids)]

    yyyy_mm_dd = f"{date_yyyymmdd[:4]}-{date_yyyymmdd[4:6]}-{date_yyyymmdd[6:]}"
    output_dir = os.path.join(os.path.dirname(__file__), "..", "site")
    os.makedirs(output_dir, exist_ok=True)

    out_xlsx = os.path.join(output_dir, f"Today_PoohPoints_SEC_ByOwner_{yyyy_mm_dd}.xlsx")
    out_players_html = os.path.join(output_dir, "today_players.html")
    out_owners_html  = os.path.join(output_dir, "today_owners.html")

    print(f"Found {len(sec_events)} SEC-involved games for {yyyy_mm_dd}\n")

    all_rows = []
    for e in sec_events:
        event_id = str(e.get("id") or "")
        hdr = extract_event_header(e)
        home = hdr["home"]
        away = hdr["away"]
        status_line = hdr["status"]

        game_label = f"{away.get('abbr','')}@{home.get('abbr','')}"
        print(f"{game_label} — {status_line} — (event {event_id})")

        players = get_boxscore_players(event_id)
        if not players:
            print("  (No boxscore player stats published yet — try again later.)\n")
            continue

        for p in players:
            key = norm_name(p["player"])
            info = draft_map.get(key)
            if info:
                owner = info["owner"]
                started_today = "Yes" if info["started"] else "No"
            else:
                owner = "Undrafted"
                started_today = "No"

            all_rows.append({
                "date": yyyy_mm_dd,
                "game": game_label,
                "status": status_line,
                "owner": owner,
                "started_today": started_today,
                "team": p["team"],
                "player": p["player"],
                "pooh": p["pooh"],
                "pts": p["pts"],
                "reb": p["reb"],
                "ast": p["ast"],
                "stl": p["stl"],
                "blk": p["blk"],
                "to":  p["to"],
                "min": p["min"],
            })

        print(f"  Players captured: {len(players)}\n")

    owner_rank = {o: i for i, o in enumerate(owner_order)}

    def sort_key(r):
        o = r["owner"]
        oidx = owner_rank.get(o, 10_000 if o == "Undrafted" else 9_000)
        starter_rank = 0 if r["started_today"] == "Yes" else 1
        return (oidx, o, starter_rank, -r["pooh"], r["player"])

    all_rows.sort(key=sort_key)

    totals: Dict[str, Dict[str, int]] = {}
    for r in all_rows:
        owner = r["owner"]
        if owner == "Undrafted":
            continue

        if owner not in totals:
            totals[owner] = {"starter_pooh_total": 0, "starters_count_so_far": 0}

        if r["started_today"] == "Yes":
            totals[owner]["starter_pooh_total"] += int(r["pooh"])
            totals[owner]["starters_count_so_far"] += 1

    owner_totals_rows = [{"owner": o, **vals} for o, vals in totals.items()]
    owner_totals_rows.sort(key=lambda x: x["starter_pooh_total"], reverse=True)

    write_xlsx(all_rows, owner_totals_rows, out_xlsx)
    write_html_tables(all_rows, owner_totals_rows, out_players_html, out_owners_html, yyyy_mm_dd)

    print(f"Wrote: {out_players_html}")
    print(f"Wrote: {out_owners_html}")
    print(f"Wrote: {out_xlsx}")


if __name__ == "__main__":
    main()
