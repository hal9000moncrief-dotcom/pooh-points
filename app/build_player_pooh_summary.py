import os
import re
import glob
import math
import html
from datetime import datetime
from typing import Dict, Any, List, Tuple, Optional

from bs4 import BeautifulSoup
from openpyxl import load_workbook


REPO_ROOT = os.path.join(os.path.dirname(__file__), "..")
APP_DIR = os.path.join(REPO_ROOT, "app")
DOCS_DIR = os.path.join(REPO_ROOT, "docs")

ROSTERS_XLSX = os.path.join(APP_DIR, "Rosters.xlsx")
OUT_HTML = os.path.join(DOCS_DIR, "Player_Pooh_Summary.html")

FINAL_PLAYERS_GLOB = os.path.join(DOCS_DIR, "Final_Players_PD*.html")


# ----------------------------
# Name normalization (match your existing logic closely)
# ----------------------------
def norm_name(name: str) -> str:
    s = (name or "").lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\b(jr|sr|ii|iii|iv)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def safe_float(v) -> float:
    try:
        if v is None:
            return 0.0
        s = str(v).strip()
        if s == "" or s == "--":
            return 0.0
        return float(s)
    except Exception:
        return 0.0


def safe_int(v) -> int:
    try:
        if v is None:
            return 0
        s = str(v).strip()
        if s == "" or s == "--":
            return 0
        return int(float(s))
    except Exception:
        return 0


def parse_pd_num_from_filename(path: str) -> Optional[int]:
    m = re.search(r"Final_Players_PD(\d+)\.html$", os.path.basename(path), re.IGNORECASE)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


# ----------------------------
# Rosters.xlsx loader (header-flexible)
# ----------------------------
def load_rosters(xlsx_path: str) -> Dict[str, Dict[str, Any]]:
    """
    Returns dict keyed by normalized player name, values include:
      pid, team_name, cost, name, team, height, weight, class, position
    We try to match headers flexibly.
    """
    if not os.path.isfile(xlsx_path):
        raise RuntimeError(f"Missing Rosters file: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Map headers
    header_map: Dict[str, int] = {}
    for cell in ws[1]:
        h = (str(cell.value).strip() if cell.value is not None else "")
        if h:
            header_map[h.lower()] = cell.column

    def col(*names) -> Optional[int]:
        for n in names:
            if n.lower() in header_map:
                return header_map[n.lower()]
        return None

    c_name = col("name", "player", "player name")
    c_pid = col("pid", "playerid", "player id")
    c_team_name = col("team name", "teamname", "owner team", "fantasy team", "team")
    c_cost = col("cost", "auction cost", "price")
    c_school = col("team", "school", "college", "school team")
    c_height = col("height")
    c_weight = col("weight")
    c_class = col("class", "yr", "year")
    c_position = col("position", "pos")

    if c_name is None:
        raise RuntimeError("Rosters.xlsx must have a Name column (e.g., 'Name').")

    out: Dict[str, Dict[str, Any]] = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=c_name).value
        if name is None or str(name).strip() == "":
            continue

        rec = {
            "pid": ws.cell(row=r, column=c_pid).value if c_pid else "",
            "team_name": ws.cell(row=r, column=c_team_name).value if c_team_name else "",
            "cost": ws.cell(row=r, column=c_cost).value if c_cost else "",
            "name": str(name).strip(),
            "team": ws.cell(row=r, column=c_school).value if c_school else "",
            "height": ws.cell(row=r, column=c_height).value if c_height else "",
            "weight": ws.cell(row=r, column=c_weight).value if c_weight else "",
            "class": ws.cell(row=r, column=c_class).value if c_class else "",
            "position": ws.cell(row=r, column=c_position).value if c_position else "",
        }
        out[norm_name(rec["name"])] = rec

    return out


# ----------------------------
# Parse Final_Players_PD*.html
# ----------------------------
def parse_final_players_html(path: str) -> List[Dict[str, Any]]:
    """
    Extract rows with fields at least:
      player, pooh, pts, reb, ast, stl, blk, to, min
    Your Final_Players pages come from write_html_tables() with columns including these.
    """
    with open(path, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f.read(), "lxml")

    table = soup.find("table")
    if not table:
        return []

    # Read header row
    header_cells = table.find("thead").find_all("th") if table.find("thead") else table.find_all("th")
    headers = [h.get_text(strip=True) for h in header_cells]

    # Build column index
    idx = {h: i for i, h in enumerate(headers)}

    required = ["player", "pooh", "pts", "reb", "ast", "stl", "blk", "to", "min"]
    # If any are missing, return empty to avoid wrong parsing
    for r in required:
        if r not in idx:
            return []

    rows = []
    body = table.find("tbody") or table
    for tr in body.find_all("tr"):
        tds = tr.find_all(["td", "th"])
        if not tds:
            continue
        vals = [td.get_text(strip=True) for td in tds]
        if len(vals) < len(headers):
            # pad
            vals += [""] * (len(headers) - len(vals))

        player = vals[idx["player"]]
        if not player:
            continue

        rows.append({
            "player": player,
            "pooh": safe_int(vals[idx["pooh"]]),
            "pts": safe_int(vals[idx["pts"]]),
            "reb": safe_int(vals[idx["reb"]]),
            "ast": safe_int(vals[idx["ast"]]),
            "stl": safe_int(vals[idx["stl"]]),
            "blk": safe_int(vals[idx["blk"]]),
            "to": safe_int(vals[idx["to"]]),
            "min": safe_float(vals[idx["min"]]),
        })

    return rows


# ----------------------------
# HTML writer
# ----------------------------
def write_player_summary_html(rows: List[Dict[str, Any]], out_path: str, title: str):
    def esc(x):
        return html.escape("" if x is None else str(x))

    pd_cols = [str(i) for i in range(1, 20)]

    # Columns per your screenshot (plus PD1..PD19 and per-game stats)
    cols = (
        ["#", "PID", "Team Name", "Cost", "Name", "Team", "Height", "Weight", "Class", "Position", "Min/G", "Avg", "Total"]
        + pd_cols
        + ["PPG", "R/G", "A/G", "B/G", "S/G", "T/G"]
    )

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("<!doctype html><html><head><meta charset='utf-8'>")
        f.write(f"<title>{esc(title)}</title>")
        f.write(
            "<style>"
            "body{font-family:Arial;background:#ffffff}"
            ".wrap{width:1400px;max-width:98vw;margin:18px auto;border:3px solid #000;background:#FFFFCC;padding:10px}"
            "h2{text-align:center;margin:6px 0 12px 0}"
            "table{border-collapse:collapse;width:100%;background:#fff}"
            "th,td{border:1px solid #000;padding:6px 6px;text-align:center;font-size:12pt;white-space:nowrap}"
            "th{background:#c0c0c0}"
            ".left{text-align:left}"
            ".pid,.cost{font-weight:700}"
            ".small{font-size:11pt}"
            "</style>"
        )
        f.write("</head><body><div class='wrap'>")
        f.write(f"<h2>{esc(title)}</h2>")
        f.write(f"<div class='small'><b>Last updated:</b> {esc(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}</div>")
        f.write("<br>")

        f.write("<table><thead><tr>")
        for c in cols:
            f.write(f"<th>{esc(c)}</th>")
        f.write("</tr></thead><tbody>")

        for r in rows:
            f.write("<tr>")
            for c in cols:
                v = r.get(c, "")
                cls = ""
                if c in ("Name", "Team Name"):
                    cls = " class='left'"
                if c in ("PID",):
                    cls = " class='pid'"
                if c in ("Cost",):
                    cls = " class='cost'"
                f.write(f"<td{cls}>{esc(v)}</td>")
            f.write("</tr>")

        f.write("</tbody></table></div></body></html>")


def main():
    rosters = load_rosters(ROSTERS_XLSX)

    final_files = sorted(glob.glob(FINAL_PLAYERS_GLOB), key=lambda p: parse_pd_num_from_filename(p) or 9999)
    if not final_files:
        raise SystemExit("No Final_Players_PD*.html files found in docs/. Run a Final Run first.")

    # Aggregation structure per normalized player
    agg: Dict[str, Dict[str, Any]] = {}

    for fp in final_files:
        pd = parse_pd_num_from_filename(fp)
        if not pd or pd < 1 or pd > 19:
            continue

        rows = parse_final_players_html(fp)
        for rr in rows:
            key = norm_name(rr["player"])
            roster = rosters.get(key)

            # Skip players not in Rosters.xlsx (keeps page clean)
            if roster is None:
                continue

            if key not in agg:
                # initialize
                agg[key] = {
                    "PID": roster.get("pid", ""),
                    "Team Name": roster.get("team_name", ""),
                    "Cost": roster.get("cost", ""),
                    "Name": roster.get("name", rr["player"]),
                    "Team": roster.get("team", rr.get("team", "")),
                    "Height": roster.get("height", ""),
                    "Weight": roster.get("weight", ""),
                    "Class": roster.get("class", ""),
                    "Position": roster.get("position", ""),
                    # PD columns:
                    **{str(i): "" for i in range(1, 20)},
                    # totals:
                    "_games": 0,
                    "_pooh_total": 0,
                    "_min_total": 0.0,
                    "_pts_total": 0,
                    "_reb_total": 0,
                    "_ast_total": 0,
                    "_blk_total": 0,
                    "_stl_total": 0,
                    "_to_total": 0,
                }

            # PD value = pooh for that PD (if multiple rows exist somehow, sum)
            current_pd_val = safe_int(agg[key][str(pd)]) if str(agg[key][str(pd)]).strip() != "" else 0
            new_pd_val = current_pd_val + int(rr["pooh"])
            agg[key][str(pd)] = new_pd_val

            # per-game totals
            agg[key]["_games"] += 1
            agg[key]["_pooh_total"] += int(rr["pooh"])
            agg[key]["_min_total"] += float(rr["min"])
            agg[key]["_pts_total"] += int(rr["pts"])
            agg[key]["_reb_total"] += int(rr["reb"])
            agg[key]["_ast_total"] += int(rr["ast"])
            agg[key]["_blk_total"] += int(rr["blk"])
            agg[key]["_stl_total"] += int(rr["stl"])
            agg[key]["_to_total"] += int(rr["to"])

    # Build output rows
    out_rows = []
    for key, a in agg.items():
        g = max(1, int(a["_games"]))

        total = int(a["_pooh_total"])
        avg = total / g
        min_g = a["_min_total"] / g

        row = {
            "PID": a["PID"],
            "Team Name": a["Team Name"],
            "Cost": a["Cost"],
            "Name": a["Name"],
            "Team": a["Team"],
            "Height": a["Height"],
            "Weight": a["Weight"],
            "Class": a["Class"],
            "Position": a["Position"],
            "Min/G": f"{min_g:.1f}",
            "Avg": f"{avg:.2f}",
            "Total": total,
            "PPG": f"{a['_pts_total']/g:.1f}",
            "R/G": f"{a['_reb_total']/g:.1f}",
            "A/G": f"{a['_ast_total']/g:.1f}",
            "B/G": f"{a['_blk_total']/g:.1f}",
            "S/G": f"{a['_stl_total']/g:.1f}",
            "T/G": f"{a['_to_total']/g:.1f}",
        }

        # PD1..PD19 (blank if 0)
        for i in range(1, 20):
            v = a[str(i)]
            if v == "" or safe_int(v) == 0:
                row[str(i)] = ""
            else:
                row[str(i)] = int(v)

        out_rows.append(row)

    # Sort: Avg desc, Total desc, Name asc
    def sort_key(r):
        return (-float(r["Avg"]), -int(r["Total"]), str(r["Name"]).lower())

    out_rows.sort(key=sort_key)

    # Add rank column "#"
    for i, r in enumerate(out_rows, start=1):
        r["#"] = i

    title = "Player Pooh Summary — 2025–2026 Regular Season"
    write_player_summary_html(out_rows, OUT_HTML, title)
    print(f"Wrote: {OUT_HTML}")


if __name__ == "__main__":
    main()
