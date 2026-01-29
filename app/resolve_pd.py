import re
from datetime import datetime, date, timedelta
from openpyxl import load_workbook

def norm_to_mmddyyyy(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%m%d%Y")
    if isinstance(v, date):
        return v.strftime("%m%d%Y")
    if isinstance(v, (int, float)):
        iv = int(v)
        # Excel serial date
        if 30000 <= iv <= 80000:
            d = date(1899, 12, 30) + timedelta(days=iv)
            return d.strftime("%m%d%Y")
        s = str(iv)
        return s.zfill(8) if len(s) <= 8 else None

    s = str(v).strip()

    m = re.match(r"^(\d{8})(?:\.0)?$", s)
    if m:
        return m.group(1)

    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        mo, da, yr = m.groups()
        return f"{int(mo):02d}{int(da):02d}{int(yr):04d}"

    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", s)
    if m:
        yr, mo, da = m.groups()
        return f"{int(mo):02d}{int(da):02d}{int(yr):04d}"

    m = re.match(r"^(\d{6,8})(?:\.0)?$", s)
    if m:
        return m.group(1).zfill(8)

    return None

def resolve_pd(xlsx_path: str, yyyymmdd: str) -> str:
    yyyymmdd = (yyyymmdd or "").strip()
    if not re.fullmatch(r"\d{8}", yyyymmdd):
        raise SystemExit(f"ERROR: date must be YYYYMMDD (8 digits). Got: {yyyymmdd}")

    yyyy = yyyymmdd[0:4]
    mm   = yyyymmdd[4:6]
    dd   = yyyymmdd[6:8]
    target_mmddyyyy = f"{mm}{dd}{yyyy}"

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    found = None
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        left = norm_to_mmddyyyy(row[0])
        if left == target_mmddyyyy:
            found = row[1]
            break

    if found is None:
        raise SystemExit(
            f"ERROR: Date {yyyymmdd} (MMDDYYYY={target_mmddyyyy}) is not in {xlsx_path}."
        )

    pdnum = str(found).strip().replace(".0", "")
    if not pdnum.isdigit():
        raise SystemExit(f"ERROR: PD value for {target_mmddyyyy} is not numeric: {found}")

    return f"PD{int(pdnum)}"

if __name__ == "__main__":
    # args: PD.xlsx path, yyyymmdd
    import sys
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python app/resolve_pd.py app/PD.xlsx YYYYMMDD")
    pd = resolve_pd(sys.argv[1], sys.argv[2])
    print(pd)
    with open("pd_resolved.txt", "w", encoding="utf-8") as f:
        f.write(pd)
